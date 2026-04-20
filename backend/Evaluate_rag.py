"""
RAG 评估系统
============
包含三部分：
  1. 检索评估：准确率、召回率（有/无 HyDE 对比）
  2. 回答质量：LLM 自动打分（1-5分）
  3. 结果导出：Excel 报告 + 控制台摘要

使用方法：
  python evaluate_rag.py
  python evaluate_rag.py --mode retrieval   # 只跑检索评估
  python evaluate_rag.py --mode answer      # 只跑回答质量评估
  python evaluate_rag.py --mode all         # 全部跑（默认）
"""

import argparse
import json
import os
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from dotenv import load_dotenv
from openai import OpenAI

load_dotenv(Path(__file__).parent / ".env")

# ============================================================
# 评估数据集
# 格式：{ question, relevant_keywords, reference_answer }
#   - relevant_keywords：答案里应该包含的关键词（用于判断文档相关性）
#   - reference_answer：标准参考答案（用于 LLM 评分对比）
# ============================================================

EVAL_DATASET = [
    # ---- 病害类 ----
    {
        "id": "Q01",
        "question": "脐橙黄龙病的症状是什么？",
        "relevant_keywords": ["黄龙病", "黄化", "斑驳", "叶片"],
        "reference_answer": "黄龙病主要症状为叶片出现黄绿相间的斑驳，果实变小、味酸，最终导致树体衰退死亡。",
        "category": "病害",
    },
    {
        "id": "Q02",
        "question": "溃疡病如何防治？",
        "relevant_keywords": ["溃疡病", "铜制剂", "波尔多液", "防治"],
        "reference_answer": "溃疡病防治需在嫩梢期喷施铜制剂，结合剪除病枝，避免机械伤口。",
        "category": "病害",
    },
    {
        "id": "Q03",
        "question": "脐橙炭疽病怎么治？",
        "relevant_keywords": ["炭疽病", "杀菌剂", "多菌灵", "甲基托布津"],
        "reference_answer": "炭疽病可用多菌灵或甲基托布津喷雾防治，注意增强树势提高抗病能力。",
        "category": "病害",
    },
    {
        "id": "Q04",
        "question": "脐橙根腐病的原因和处理方法",
        "relevant_keywords": ["根腐病", "排水", "土壤", "根系"],
        "reference_answer": "根腐病多因积水导致，需改善排水条件，挖除腐烂根系，用杀菌剂灌根处理。",
        "category": "病害",
    },
    # ---- 虫害类 ----
    {
        "id": "Q05",
        "question": "柑橘木虱如何防治？",
        "relevant_keywords": ["木虱", "黄龙病", "嫩梢", "啶虫脒"],
        "reference_answer": "木虱是黄龙病的传播媒介，需在嫩梢期喷施啶虫脒、吡虫啉等药剂，统防统治。",
        "category": "虫害",
    },
    {
        "id": "Q06",
        "question": "红蜘蛛危害脐橙怎么办？",
        "relevant_keywords": ["红蜘蛛", "螨类", "杀螨剂", "阿维菌素"],
        "reference_answer": "红蜘蛛高发期喷施阿维菌素或哒螨灵，注意轮换用药避免产生抗性。",
        "category": "虫害",
    },
    {
        "id": "Q07",
        "question": "脐橙介壳虫怎么防治？",
        "relevant_keywords": ["介壳虫", "蚧壳虫", "矿物油", "若虫"],
        "reference_answer": "防治介壳虫宜在若虫孵化期喷施矿物油乳剂或噻嗪酮，抓住孵化盛期用药。",
        "category": "虫害",
    },
    {
        "id": "Q08",
        "question": "潜叶蛾如何防治？",
        "relevant_keywords": ["潜叶蛾", "嫩梢", "阿维菌素", "叶面"],
        "reference_answer": "潜叶蛾主要危害嫩梢，在嫩芽萌发至1厘米时开始喷药，阿维菌素效果好。",
        "category": "虫害",
    },
    # ---- 施肥类 ----
    {
        "id": "Q09",
        "question": "脐橙一年需要施几次肥？什么时候施？",
        "relevant_keywords": ["施肥", "次数", "花前肥", "壮果肥", "采后肥"],
        "reference_answer": "脐橙一般一年施3-4次肥：花前肥（2-3月）、壮果肥（6-7月）、采后肥（11月后）。",
        "category": "施肥",
    },
    {
        "id": "Q10",
        "question": "脐橙缺钙有什么表现？怎么补充？",
        "relevant_keywords": ["缺钙", "裂果", "石灰", "钙肥"],
        "reference_answer": "缺钙表现为果实裂果、叶片畸形，可喷施氯化钙或叶面钙肥，土壤施石灰调节pH。",
        "category": "施肥",
    },
    {
        "id": "Q11",
        "question": "脐橙叶片发黄缺什么肥？",
        "relevant_keywords": ["黄化", "缺氮", "缺铁", "缺镁", "叶面肥"],
        "reference_answer": "叶片整体黄化多缺氮，叶脉绿叶肉黄多缺铁或缺镁，需针对性补充相应元素。",
        "category": "施肥",
    },
    # ---- 栽培管理类 ----
    {
        "id": "Q12",
        "question": "脐橙什么时候修剪最合适？",
        "relevant_keywords": ["修剪", "采果后", "春季", "疏枝"],
        "reference_answer": "脐橙主要在采果后至春芽萌发前修剪，剪除枯枝、病枝、交叉枝，改善通风透光。",
        "category": "栽培",
    },
    {
        "id": "Q13",
        "question": "脐橙如何保花保果？",
        "relevant_keywords": ["保花保果", "环割", "赤霉素", "落花落果"],
        "reference_answer": "保花保果可在花期和幼果期喷施赤霉素，配合合理疏花疏果，减少落果。",
        "category": "栽培",
    },
    {
        "id": "Q14",
        "question": "脐橙种植密度多少合适？",
        "relevant_keywords": ["种植密度", "株距", "行距", "亩株数"],
        "reference_answer": "一般采用3米×4米或4米×5米株行距，每亩种植33-55株，根据地形和品种调整。",
        "category": "栽培",
    },
    {
        "id": "Q15",
        "question": "脐橙果实套袋有什么好处？",
        "relevant_keywords": ["套袋", "果实", "外观", "病虫害"],
        "reference_answer": "套袋可改善果实外观色泽、减少病虫危害和农药残留，提高商品价值。",
        "category": "栽培",
    },
    # ---- 采后处理类 ----
    {
        "id": "Q16",
        "question": "脐橙采摘后如何储存？",
        "relevant_keywords": ["储存", "贮藏", "温度", "湿度", "保鲜"],
        "reference_answer": "脐橙采后应在阴凉通风处预冷，贮藏温度5-8℃，湿度85-90%，可储存2-3个月。",
        "category": "采后",
    },
    {
        "id": "Q17",
        "question": "脐橙采后如何处理才能延长保质期？",
        "relevant_keywords": ["采后", "防腐", "打蜡", "保鲜剂"],
        "reference_answer": "采后用抑霉唑等防腐剂处理，配合打蜡保鲜，低温冷链运输可有效延长保质期。",
        "category": "采后",
    },
    # ---- 品种类 ----
    {
        "id": "Q18",
        "question": "赣南地区适合种植什么脐橙品种？",
        "relevant_keywords": ["品种", "赣南", "纽荷尔", "朋娜"],
        "reference_answer": "赣南地区主栽品种为纽荷尔脐橙，其次有朋娜、奈维林娜等，纽荷尔品质最优。",
        "category": "品种",
    },
    {
        "id": "Q19",
        "question": "纽荷尔脐橙和朋娜脐橙有什么区别？",
        "relevant_keywords": ["纽荷尔", "朋娜", "品种", "果实特征"],
        "reference_answer": "纽荷尔果形椭圆、皮薄光滑、化渣性好；朋娜果形较圆、肉质细嫩，成熟期略早。",
        "category": "品种",
    },
    # ---- 气象类 ----
    {
        "id": "Q20",
        "question": "脐橙低温冻害怎么预防和救治？",
        "relevant_keywords": ["冻害", "低温", "防寒", "熏烟"],
        "reference_answer": "冬季寒潮前可熏烟、覆盖保温，受冻后及时剪除枯枝，加强施肥促进恢复。",
        "category": "气象",
    },
]


# ============================================================
# 工具函数
# ============================================================

def get_llm_client() -> OpenAI:
    return OpenAI(
        api_key=os.getenv("DASHSCOPE_API_KEY"),
        base_url=os.getenv("OLLAMA_BASE_URL", "https://dashscope.aliyuncs.com/compatible-mode/v1"),
    )

def get_model() -> str:
    return os.getenv("OLLAMA_MODEL", "qwen-max")


# ============================================================
# Part 1：检索评估
# ============================================================

def evaluate_retrieval(kb, use_enhanced: bool = True) -> Dict:
    """
    评估检索准确率和召回率。

    准确率（Precision）= 检索到的文档中，相关文档的比例
    召回率（Recall）   = 标注相关关键词中，被检索文档覆盖的比例

    参数:
        kb: VectorKnowledgeBase 实例
        use_enhanced: True=使用 HyDE+多查询, False=使用普通检索
    """
    method_name = "增强检索（HyDE+多查询）" if use_enhanced else "普通检索"
    print(f"\n{'='*60}")
    print(f"检索评估 - {method_name}")
    print(f"{'='*60}")

    results = []
    total_precision = 0.0
    total_recall = 0.0

    for item in EVAL_DATASET:
        question = item["question"]
        keywords = item["relevant_keywords"]

        # 执行检索
        try:
            if use_enhanced:
                docs = kb.search_enhanced(question)
            else:
                docs = kb.search_docs(question)
        except Exception as e:
            print(f"  ❌ {item['id']} 检索失败: {e}")
            docs = []

        # 计算每个文档是否相关（包含任意关键词即视为相关）
        relevant_docs = []
        for doc in docs:
            content = doc.page_content
            is_relevant = any(kw in content for kw in keywords)
            relevant_docs.append(is_relevant)

        retrieved_count = len(docs)
        relevant_count = sum(relevant_docs)

        # 准确率：检索到的文档中有多少是相关的
        precision = relevant_count / retrieved_count if retrieved_count > 0 else 0.0

        # 召回率：关键词中有多少被覆盖
        all_content = " ".join(doc.page_content for doc in docs)
        covered_keywords = [kw for kw in keywords if kw in all_content]
        recall = len(covered_keywords) / len(keywords) if keywords else 0.0

        total_precision += precision
        total_recall += recall

        result = {
            "id": item["id"],
            "question": question,
            "category": item["category"],
            "retrieved_count": retrieved_count,
            "relevant_count": relevant_count,
            "precision": round(precision, 3),
            "recall": round(recall, 3),
            "covered_keywords": covered_keywords,
            "missing_keywords": [kw for kw in keywords if kw not in all_content],
        }
        results.append(result)

        status = "✅" if precision >= 0.5 and recall >= 0.5 else "⚠️"
        print(f"  {status} {item['id']} [{item['category']}] 准确率:{precision:.2f} 召回率:{recall:.2f} | {question[:20]}...")

    avg_precision = total_precision / len(EVAL_DATASET)
    avg_recall = total_recall / len(EVAL_DATASET)
    f1 = 2 * avg_precision * avg_recall / (avg_precision + avg_recall) if (avg_precision + avg_recall) > 0 else 0

    summary = {
        "method": method_name,
        "avg_precision": round(avg_precision, 3),
        "avg_recall": round(avg_recall, 3),
        "f1_score": round(f1, 3),
        "details": results,
    }

    print(f"\n  📊 平均准确率: {avg_precision:.3f}")
    print(f"  📊 平均召回率: {avg_recall:.3f}")
    print(f"  📊 F1 Score:   {f1:.3f}")

    return summary


# ============================================================
# Part 2：回答质量评估（LLM 自动打分）
# ============================================================

SCORE_PROMPT = """你是一个农业知识问答系统的评估专家。
请对以下回答进行评分，评分标准如下：

5分：回答完整准确，覆盖核心要点，表达清晰可执行
4分：回答基本准确，覆盖主要要点，有少量遗漏
3分：回答部分正确，核心要点有遗漏，但无明显错误
2分：回答有明显遗漏或模糊，实用性差
1分：回答错误或与问题无关

问题：{question}

参考答案：{reference}

系统回答：{answer}

请只输出一个1-5的整数评分，不要任何解释。
评分："""


def score_answer(llm: OpenAI, model: str, question: str, reference: str, answer: str) -> int:
    """用 LLM 对回答打分，返回 1-5 的整数"""
    prompt = SCORE_PROMPT.format(
        question=question,
        reference=reference,
        answer=answer,
    )
    try:
        response = llm.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": prompt}],
            temperature=0.0,
            max_tokens=10,
        )
        text = response.choices[0].message.content.strip()
        score = int(text[0])  # 取第一个字符
        return max(1, min(5, score))  # 限制在 1-5 范围
    except Exception:
        return 0  # 打分失败返回 0


def evaluate_answer_quality(rag_service, use_enhanced: bool = True) -> Dict:
    """
    评估回答质量，用 LLM 自动打分。
    """
    method_name = "增强检索（HyDE+多查询）" if use_enhanced else "普通检索"
    print(f"\n{'='*60}")
    print(f"回答质量评估 - {method_name}")
    print(f"{'='*60}")

    llm = get_llm_client()
    model = get_model()
    results = []
    total_score = 0
    valid_count = 0

    for item in EVAL_DATASET:
        question = item["question"]
        reference = item["reference_answer"]

        # 获取系统回答
        try:
            # 临时切换检索方式
            kb = rag_service.agent.kb
            if use_enhanced:
                # 让 KnowledgeQATool 用增强检索
                result = rag_service.ask_question(question)
            else:
                # 临时 monkey-patch 为普通检索
                original = kb.search_enhanced
                kb.search_enhanced = kb.search_docs
                result = rag_service.ask_question(question)
                kb.search_enhanced = original

            answer = result.get("answer", "")
        except Exception as e:
            answer = f"[获取回答失败: {e}]"

        # LLM 打分
        score = score_answer(llm, model, question, reference, answer)

        if score > 0:
            total_score += score
            valid_count += 1

        result_item = {
            "id": item["id"],
            "question": question,
            "category": item["category"],
            "reference": reference,
            "answer": answer,
            "score": score,
        }
        results.append(result_item)

        stars = "⭐" * score if score > 0 else "❌"
        print(f"  {stars}({score}分) {item['id']} [{item['category']}] {question[:25]}...")

        time.sleep(0.5)  # 避免触发 API 限速

    avg_score = total_score / valid_count if valid_count > 0 else 0

    summary = {
        "method": method_name,
        "avg_score": round(avg_score, 2),
        "score_distribution": {
            str(i): sum(1 for r in results if r["score"] == i)
            for i in range(1, 6)
        },
        "details": results,
    }

    print(f"\n  📊 平均分: {avg_score:.2f} / 5.00")
    print(f"  📊 分数分布: " + " | ".join(
        f"{k}分:{v}题" for k, v in summary["score_distribution"].items()
    ))

    return summary


# ============================================================
# Part 3：对比实验汇总
# ============================================================

def run_comparison(rag_service) -> Dict:
    """运行有/无增强检索的对比实验"""
    kb = rag_service.agent.kb

    print("\n🔍 开始对比实验：普通检索 vs 增强检索（HyDE+多查询）")
    print("预计耗时：5-10 分钟（需要多次 API 调用）\n")

    # 检索评估对比
    retrieval_normal = evaluate_retrieval(kb, use_enhanced=False)
    retrieval_enhanced = evaluate_retrieval(kb, use_enhanced=True)

    # 回答质量对比
    answer_normal = evaluate_answer_quality(rag_service, use_enhanced=False)
    answer_enhanced = evaluate_answer_quality(rag_service, use_enhanced=True)

    # 汇总对比
    comparison = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "retrieval": {
            "normal": {
                "precision": retrieval_normal["avg_precision"],
                "recall": retrieval_normal["avg_recall"],
                "f1": retrieval_normal["f1_score"],
            },
            "enhanced": {
                "precision": retrieval_enhanced["avg_precision"],
                "recall": retrieval_enhanced["avg_recall"],
                "f1": retrieval_enhanced["f1_score"],
            },
            "improvement": {
                "precision": round(retrieval_enhanced["avg_precision"] - retrieval_normal["avg_precision"], 3),
                "recall": round(retrieval_enhanced["avg_recall"] - retrieval_normal["avg_recall"], 3),
                "f1": round(retrieval_enhanced["f1_score"] - retrieval_normal["f1_score"], 3),
            },
        },
        "answer_quality": {
            "normal_avg": answer_normal["avg_score"],
            "enhanced_avg": answer_enhanced["avg_score"],
            "improvement": round(answer_enhanced["avg_score"] - answer_normal["avg_score"], 2),
        },
        "details": {
            "retrieval_normal": retrieval_normal,
            "retrieval_enhanced": retrieval_enhanced,
            "answer_normal": answer_normal,
            "answer_enhanced": answer_enhanced,
        },
    }

    return comparison


def print_comparison_summary(comparison: Dict):
    """打印对比结果摘要"""
    print(f"\n{'='*60}")
    print("📋 对比实验结果汇总")
    print(f"{'='*60}")

    r = comparison["retrieval"]
    a = comparison["answer_quality"]

    print("\n【检索性能对比】")
    print(f"{'指标':<12} {'普通检索':>10} {'增强检索':>10} {'提升':>8}")
    print("-" * 44)
    print(f"{'准确率':<12} {r['normal']['precision']:>10.3f} {r['enhanced']['precision']:>10.3f} {r['improvement']['precision']:>+8.3f}")
    print(f"{'召回率':<12} {r['normal']['recall']:>10.3f} {r['enhanced']['recall']:>10.3f} {r['improvement']['recall']:>+8.3f}")
    print(f"{'F1 Score':<12} {r['normal']['f1']:>10.3f} {r['enhanced']['f1']:>10.3f} {r['improvement']['f1']:>+8.3f}")

    print("\n【回答质量对比（满分5分）】")
    print(f"{'普通检索平均分':<16} {a['normal_avg']:.2f}")
    print(f"{'增强检索平均分':<16} {a['enhanced_avg']:.2f}")
    print(f"{'提升':<16} {a['improvement']:+.2f}")


# ============================================================
# Part 4：导出 Excel 报告
# ============================================================

def export_excel(comparison: Dict, output_path: str = "rag_evaluation_report.xlsx"):
    """导出 Excel 评估报告"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("\n⚠️  未安装 openpyxl，跳过 Excel 导出。安装命令：pip install openpyxl")
        # 改为导出 JSON
        json_path = output_path.replace(".xlsx", ".json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(comparison, f, ensure_ascii=False, indent=2)
        print(f"✅ 已导出 JSON 报告：{json_path}")
        return

    wb = openpyxl.Workbook()

    # ---- Sheet 1：对比摘要 ----
    ws1 = wb.active
    ws1.title = "对比摘要"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill("solid", fgColor="FFA500")  # 橙色

    ws1.append(["脐橙 RAG 系统评估报告"])
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append([f"生成时间：{comparison['timestamp']}"])
    ws1.append([])

    ws1.append(["检索性能对比", "", "", ""])
    ws1.append(["指标", "普通检索", "增强检索（HyDE+多查询）", "提升"])
    for cell in ws1[5]:
        cell.font = header_font
        cell.fill = header_fill

    r = comparison["retrieval"]
    ws1.append(["准确率（Precision）", r["normal"]["precision"], r["enhanced"]["precision"], f"{r['improvement']['precision']:+.3f}"])
    ws1.append(["召回率（Recall）",    r["normal"]["recall"],    r["enhanced"]["recall"],    f"{r['improvement']['recall']:+.3f}"])
    ws1.append(["F1 Score",           r["normal"]["f1"],        r["enhanced"]["f1"],        f"{r['improvement']['f1']:+.3f}"])

    ws1.append([])
    ws1.append(["回答质量对比（满分5分）", "", "", ""])
    ws1.append(["方法", "平均分", "", ""])
    a = comparison["answer_quality"]
    ws1.append(["普通检索", a["normal_avg"]])
    ws1.append(["增强检索", a["enhanced_avg"]])
    ws1.append(["提升", f"{a['improvement']:+.2f}"])

    for col in range(1, 5):
        ws1.column_dimensions[get_column_letter(col)].width = 28

    # ---- Sheet 2：检索详细结果 ----
    ws2 = wb.create_sheet("检索详细结果")
    headers = ["题号", "类别", "问题", "普通-准确率", "普通-召回率", "增强-准确率", "增强-召回率", "增强-缺失关键词"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    normal_details = {d["id"]: d for d in comparison["details"]["retrieval_normal"]["details"]}
    enhanced_details = {d["id"]: d for d in comparison["details"]["retrieval_enhanced"]["details"]}

    for item in EVAL_DATASET:
        qid = item["id"]
        nd = normal_details.get(qid, {})
        ed = enhanced_details.get(qid, {})
        ws2.append([
            qid,
            item["category"],
            item["question"],
            nd.get("precision", ""),
            nd.get("recall", ""),
            ed.get("precision", ""),
            ed.get("recall", ""),
            "、".join(ed.get("missing_keywords", [])),
        ])

    for col in [3]:
        ws2.column_dimensions[get_column_letter(col)].width = 35
    for col in [4, 5, 6, 7]:
        ws2.column_dimensions[get_column_letter(col)].width = 14

    # ---- Sheet 3：回答质量详细结果 ----
    ws3 = wb.create_sheet("回答质量详细结果")
    headers3 = ["题号", "类别", "问题", "参考答案", "普通检索回答", "普通分数", "增强检索回答", "增强分数"]
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill

    normal_ans = {d["id"]: d for d in comparison["details"]["answer_normal"]["details"]}
    enhanced_ans = {d["id"]: d for d in comparison["details"]["answer_enhanced"]["details"]}

    for item in EVAL_DATASET:
        qid = item["id"]
        na = normal_ans.get(qid, {})
        ea = enhanced_ans.get(qid, {})
        row = ws3.append([
            qid,
            item["category"],
            item["question"],
            item["reference_answer"],
            na.get("answer", ""),
            na.get("score", ""),
            ea.get("answer", ""),
            ea.get("score", ""),
        ])

    for col in [3, 4, 5, 7]:
        ws3.column_dimensions[get_column_letter(col)].width = 40
    for col in [6, 8]:
        ws3.column_dimensions[get_column_letter(col)].width = 10

    # 设置所有 sheet 自动换行
    for ws in [ws2, ws3]:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(output_path)
    print(f"\n✅ Excel 报告已导出：{output_path}")


# ============================================================
# 主入口
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="RAG 系统评估工具")
    parser.add_argument(
        "--mode",
        choices=["retrieval", "answer", "all"],
        default="all",
        help="评估模式：retrieval=只评检索, answer=只评回答质量, all=全部（默认）",
    )
    parser.add_argument("--no-excel", action="store_true", help="不导出 Excel 报告")
    args = parser.parse_args()

    # 初始化 RAG 服务
    print("🚀 正在初始化 RAG 服务...")
    import sys
    sys.path.insert(0, str(Path(__file__).parent))
    from rag_service import RAGService
    rag_service = RAGService()
    kb = rag_service.agent.kb
    print("✅ RAG 服务初始化完成")

    comparison = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "retrieval": {},
        "answer_quality": {},
        "details": {},
    }

    if args.mode in ("retrieval", "all"):
        r_normal = evaluate_retrieval(kb, use_enhanced=False)
        r_enhanced = evaluate_retrieval(kb, use_enhanced=True)
        comparison["retrieval"] = {
            "normal": {"precision": r_normal["avg_precision"], "recall": r_normal["avg_recall"], "f1": r_normal["f1_score"]},
            "enhanced": {"precision": r_enhanced["avg_precision"], "recall": r_enhanced["avg_recall"], "f1": r_enhanced["f1_score"]},
            "improvement": {
                "precision": round(r_enhanced["avg_precision"] - r_normal["avg_precision"], 3),
                "recall": round(r_enhanced["avg_recall"] - r_normal["avg_recall"], 3),
                "f1": round(r_enhanced["f1_score"] - r_normal["f1_score"], 3),
            },
        }
        comparison["details"]["retrieval_normal"] = r_normal
        comparison["details"]["retrieval_enhanced"] = r_enhanced

    if args.mode in ("answer", "all"):
        a_normal = evaluate_answer_quality(rag_service, use_enhanced=False)
        a_enhanced = evaluate_answer_quality(rag_service, use_enhanced=True)
        comparison["answer_quality"] = {
            "normal_avg": a_normal["avg_score"],
            "enhanced_avg": a_enhanced["avg_score"],
            "improvement": round(a_enhanced["avg_score"] - a_normal["avg_score"], 2),
        }
        comparison["details"]["answer_normal"] = a_normal
        comparison["details"]["answer_enhanced"] = a_enhanced

    print_comparison_summary(comparison)

    if not args.no_excel:
        export_excel(comparison)

    # 同时保存 JSON
    json_path = "rag_evaluation_result.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(comparison, f, ensure_ascii=False, indent=2)
    print(f"✅ JSON 结果已保存：{json_path}")


if __name__ == "__main__":
    main()